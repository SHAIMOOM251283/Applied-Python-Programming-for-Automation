import os
import csv
from openpyxl import load_workbook

def excel_to_csv(excel_file, output_folder):
    # Load the Excel workbook
    wb = load_workbook(excel_file, read_only=True)

    # Iterate through each sheet in the workbook
    for sheet in wb.sheetnames:
        # Create a CSV filename for each sheet
        csv_filename = os.path.join(output_folder, f"{os.path.splitext(excel_file)[0]}_{sheet}.csv")

        # Open a CSV file for writing
        with open(csv_filename, 'w', newline='') as csv_file:
            csv_writer = csv.writer(csv_file)

            # Get the current sheet
            current_sheet = wb[sheet]

            # Write each row in the sheet to the CSV file
            for row in current_sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)

if __name__ == "__main__":
    # Get the current working directory
    current_directory = os.getcwd()

    # Create a new folder for CSV files
    output_folder = os.path.join(current_directory, "CSV_Output")
    os.makedirs(output_folder, exist_ok=True)

    # Iterate through all files in the current directory
    for file in os.listdir(current_directory):
        if file.endswith(".xlsx"):
            # If the file is an Excel file, convert it to CSV
            excel_to_csv(file, output_folder)

print("Conversion complete. CSV files saved in the 'CSV_Output' folder.")
