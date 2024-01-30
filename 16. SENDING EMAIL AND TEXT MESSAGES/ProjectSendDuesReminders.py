#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

#import openpyxl, smtplib, sys
# Open the spreadsheet and get the latest dues status.
#wb = openpyxl.load_workbook('duesRecords.xlsx')
#sheet = wb['Sheet1']  # Use wb['Sheet1'] instead of wb.get_sheet_by_name('Sheet1')
#lastCol = sheet.max_column  # Use max_column instead of get_highest_column
#latestMonth = sheet.cell(row=1, column=lastCol).value
# Check each member's payment status.
#unpaidMembers = {}
#for r in range(2, sheet.max_row + 1):  # Use max_row instead of get_highest_row
#    payment = sheet.cell(row=r, column=lastCol).value
#    if payment != 'paid':
#        name = sheet.cell(row=r, column=1).value
#        email = sheet.cell(row=r, column=2).value
#        unpaidMembers[name] = email
# Log in to the email account.
#smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
#smtpObj.ehlo()
#smtpObj.starttls()
#smtpObj.login('my_email_address@gmail.com', sys.argv[1])
# Send out reminder emails.
#for name, email in unpaidMembers.items():
#    body = (
#        "Subject: %s dues unpaid.\n"
#        "Dear %s,\n"
#        "Records show that you have not paid dues for %s. "
#        "Please make this payment as soon as possible. "
#        "Thank you!"
#    ) % (latestMonth, name, latestMonth)
#    print('Sending email to %s...' % email)
#    sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)
#    if sendmailStatus != {}:
#        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
#smtpObj.quit()

#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

def send_dues_reminders():
    # Open the spreadsheet and get the latest dues status.
    wb = openpyxl.load_workbook('duesRecords.xlsx')
    sheet = wb['Sheet1']
    lastCol = sheet.max_column
    latestMonth = sheet.cell(row=1, column=lastCol).value

    # Check each member's payment status.
    unpaidMembers = {}
    for r in range(2, sheet.max_row + 1):
        payment = sheet.cell(row=r, column=lastCol).value
        if payment != 'paid':
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaidMembers[name] = email

    # Log in to the email account.
    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
    smtpObj.ehlo()
    smtpObj.starttls()
    
    try:
        smtpObj.login('my_email_address@gmail.com', sys.argv[1])

        # Send out reminder emails.
        for name, email in unpaidMembers.items():
            body = (
                "Subject: %s dues unpaid.\n"
                "Dear %s,\n"
                "Records show that you have not paid dues for %s. "
                "Please make this payment as soon as possible. "
                "Thank you!"
            ) % (latestMonth, name, latestMonth)
            print('Sending email to %s...' % email)
            sendmailStatus = smtpObj.sendmail('my_email_address@gmail.com', email, body)
            if sendmailStatus != {}:
                print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
    except smtplib.SMTPException as e:
        print(f"An error occurred: {e}")
    finally:
        smtpObj.quit()

if __name__ == "__main__":
    send_dues_reminders()



