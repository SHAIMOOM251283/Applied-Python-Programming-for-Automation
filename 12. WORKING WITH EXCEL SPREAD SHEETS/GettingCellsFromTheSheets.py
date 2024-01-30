#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet['A1'])
#print(sheet['A1'].value)
#c = sheet['B1']
#print(c.value)
#print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
#print('Cell ' + c.coordinate + ' is ' + c.value)
#print(sheet['C1'].value)

import openpyxl
from openpyxl.utils import get_column_letter
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
# Convert the integer c.row to a string and c.column to a letter before concatenating
print('Row ' + str(c.row) + ', Column ' + get_column_letter(c.column) + ' is ' + str(c.value))
# Use c.coordinate directly as it already includes the column letter
print('Cell ' + c.coordinate + ' is ' + str(c.value))
print(sheet['C1'].value)
# Accessing cells using the cell method
cell_1_2 = sheet.cell(row=1, column=2)
print(cell_1_2)
value_1_2 = sheet.cell(row=1, column=2).value
print(value_1_2)
# Using a loop to print values in column B for odd-numbered rows
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet.get_highest_row())
#print(sheet.get_highest_column())

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
# Use max_row and max_column attributes to get the highest row and column
print(sheet.max_row)
print(sheet.max_column)
