import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active  # use active instead of get_active_sheet()
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')

import openpyxl
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active  # Use the active attribute
print(sheet['A3'].value)
wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.active  # Use the active attribute
print(sheet['A3'].value)
