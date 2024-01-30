#import openpyxl
#wb = openpyxl.Workbook()
#sheet = wb.get_sheet_by_name('Sheet')
#sheet['A1'] = 'Hello world!'
#print(sheet['A1'].value)

#import openpyxl
#wb = openpyxl.Workbook()
#sheet_name = 'Sheet'
#sheet = wb[sheet_name]
#sheet['A1'] = 'Hello world!'
#value = sheet.cell(row=1, column=1).value
#print(value)

import openpyxl
wb = openpyxl.Workbook()
sheet_name = 'Sheet'
sheet = wb[sheet_name]
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)
