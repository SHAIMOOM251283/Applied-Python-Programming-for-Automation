import sys
from openpyxl import load_workbook, Workbook

def invert_cells(input_filename, output_filename):
    try:
        # Load the Excel workbook
        workbook = load_workbook(input_filename)
        sheet = workbook.active

        # Invert the cells
        inverted_data = []
        for col_idx, column in enumerate(sheet.iter_cols(), start=1):
            inverted_data.append([cell.value for cell in column])

        # Create a new workbook for the inverted data
        inverted_workbook = Workbook()
        inverted_sheet = inverted_workbook.active

        # Write the inverted data to the new workbook
        for row_idx, row_data in enumerate(inverted_data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                inverted_sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save the new workbook
        inverted_workbook.save(output_filename)

        print(f"Spreadsheet inverted successfully and saved as {output_filename}")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python SpreadSheetCellInverter.py <input_filename> <output_filename>")
    else:
        input_filename = sys.argv[1]
        output_filename = sys.argv[2]
        invert_cells(input_filename, output_filename)

