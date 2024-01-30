#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#print(wb.get_sheet_names())
#sheet = wb.get_sheet_by_name('Sheet3')
#print(sheet)
#print(type(sheet))
#print(sheet.title)
#anotherSheet = wb.get_active_sheet()
#print(anotherSheet)

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
# Get sheet names
print(wb.sheetnames)
# Get a specific sheet by name
sheet = wb['Sheet3']
print(sheet)
print(type(sheet))
print(sheet.title)
# Get the active sheet
anotherSheet = wb.active
print(anotherSheet)
