import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active  # Use the active attribute
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

import openpyxl
wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active  # Use the active attribute
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')
