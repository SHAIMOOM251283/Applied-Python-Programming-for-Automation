import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam Bacon Eggs Sheet'
print(wb.sheetnames)
# Save the workbook to a file
#wb.save('spam.xlsx')

#import openpyxl
#wb = openpyxl.load_workbook('example.xlsx')
#sheet = wb.get_active_sheet()
#sheet.title = 'Spam Spam Spam'
#wb.save('example_copy.xlsx')

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')