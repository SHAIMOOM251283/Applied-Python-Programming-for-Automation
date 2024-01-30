#import sys
#from openpyxl import load_workbook
#from openpyxl.utils import get_column_letter

#def insert_blank_rows(filename, start_row, num_rows):
#    try:
        # Load the workbook
#        workbook = load_workbook(filename)
#        sheet = workbook.active

        # Insert M blank rows starting from row N
#        sheet.insert_rows(start_row, amount=num_rows)

        # Save the modified workbook
#        workbook.save(filename)
#        print(f"{num_rows} blank rows inserted starting from row {start_row}.")
#    except Exception as e:
#        print(f"Error: {e}")

#if __name__ == "__main__":
    # Check if the correct number of command line arguments are provided
#    if len(sys.argv) != 4:
#        print("Usage: python blankRowInserter.py <filename> <start_row> <num_rows>")
#        sys.exit(1)

    # Parse command line arguments
#    filename = sys.argv[1]
#    start_row = int(sys.argv[2])
#    num_rows = int(sys.argv[3])

    # Call the function to insert blank rows
#    insert_blank_rows(filename, start_row, num_rows)

import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def insert_blank_rows(filename, start_row, num_rows, output_filename):
    try:
        # Load the workbook
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Insert M blank rows starting from row N
        sheet.insert_rows(start_row, amount=num_rows)

        # Save the modified workbook with a new filename
        workbook.save(output_filename)
        print(f"{num_rows} blank rows inserted starting from row {start_row}.")
        print(f"Modified spreadsheet saved as {output_filename}.")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    # Check if the correct number of command line arguments are provided
    if len(sys.argv) != 5:
        print("Usage: python blankRowInserter.py <filename> <start_row> <num_rows> <output_filename>")
        sys.exit(1)

    # Parse command line arguments
    filename = sys.argv[1]
    start_row = int(sys.argv[2])
    num_rows = int(sys.argv[3])
    output_filename = sys.argv[4]

    # Call the function to insert blank rows and save the modified spreadsheet
    insert_blank_rows(filename, start_row, num_rows, output_filename)
