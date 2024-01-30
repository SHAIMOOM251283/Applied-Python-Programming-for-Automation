# The openpyxl.load_workbook() function is part of the openpyxl library, which is used for working with Excel files in Python. 
# This function is specifically designed to load an existing Excel workbook from a file.

# The get_sheet_names() method is not a direct method of the Workbook class in openpyxl.
# In openpyxl, the wb.active attribute of the Workbook object represents the currently active worksheet. 
# The active attribute is an instance of the Worksheet class.

# To retrieve the Worksheet object for a sheet named 'Sheet1' in openpyxl, you can use the workbook['Sheet1'] syntax.
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Get the Worksheet object for 'Sheet1'
sheet1 = workbook['Sheet1']
# Access a specific cell in 'Sheet1'
cell_value = sheet1['A1'].value
# Close the workbook when you're done
workbook.close()
# The provided code is an example of how to use openpyxl to load an existing Excel workbook named 'example.xlsx' and perform operations on it. 
# In this case, it demonstrates how to access a specific sheet ('Sheet1') and a specific cell ('A1') within that sheet to retrieve the cell's value.

# To retrieve the Worksheet object for the workbook's active sheet in openpyxl, you can use the active attribute of the Workbook object. 
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Get the active worksheet
active_sheet = workbook.active
# Access a specific cell in the active worksheet
cell_value = active_sheet['A1'].value
# Close the workbook when you're done
workbook.close()
# workbook.active gives you access to the active worksheet. 
# You can then use the active_sheet object to perform operations on the active sheet, such as accessing cells, modifying values, or performing other operations.

# To retrieve the value in cell C5 of a specific worksheet in openpyxl, you can use the following code:
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Specify the worksheet (e.g., 'Sheet1')
worksheet = workbook['Sheet1']
# Access the value in cell C5
cell_c5_value = worksheet['C5'].value
# Print or use the retrieved value
print(cell_c5_value)
# Close the workbook when you're done
workbook.close()
# The code then accesses the value in cell C5 of that specific worksheet using worksheet['C5'].value.

# To set the value in cell C5 to "Hello" in openpyxl, you can use the following code:
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Specify the worksheet (e.g., 'Sheet1')
worksheet = workbook['Sheet1']
# Set the value in cell C5 to "Hello"
worksheet['C5'] = "Hello"
# Save the changes to the workbook
workbook.save('example.xlsx')
# Close the workbook when you're done
workbook.close()
# The code sets the value in cell C5 to "Hello" using worksheet['C5'] = "Hello". 

# In openpyxl, you can retrieve the row and column indices of a cell as integers using the row and column attributes of the Cell object. 
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Specify the worksheet (e.g., 'Sheet1')
worksheet = workbook['Sheet1']
# Access a specific cell (e.g., C5)
cell = worksheet['C5']
# Retrieve the row and column as integers
row_index = cell.row
column_index = cell.column
# Print or use the retrieved values
print("Row:", row_index)
print("Column:", column_index)
# Close the workbook when you're done
workbook.close()
# In this example, cell.row and cell.column give you the row and column indices of the cell, respectively.

# There are no direct methods named get_highest_column() or get_highest_row() in openpyxl. 
# Instead, you can use the max_column and max_row attributes, respectively, to achieve similar functionality.
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Specify the worksheet (e.g., 'Sheet1')
worksheet = workbook['Sheet1']
# Retrieve the highest column and row indices
highest_column = worksheet.max_column
highest_row = worksheet.max_row
# Print or use the retrieved values
print("Highest Column:", highest_column)
print("Highest Row:", highest_row)
# Close the workbook when you're done
workbook.close()
# In this example, worksheet.max_column returns the highest column index with data, and worksheet.max_row returns the highest row index with data.

# To get the integer index for column 'M' in openpyxl, you can use the openpyxl.utils.column_index_from_string function. 
from openpyxl.utils import column_index_from_string
# Convert column letter 'M' to integer index
column_index = column_index_from_string('M')
# Print or use the retrieved value
print("Column 'M' Index:", column_index)
# In this example, column_index_from_string('M') returns the integer index corresponding to the column letter 'M'.
# Conversely, if you have an integer index and you want to get the corresponding column letter, you can use the openpyxl.utils.get_column_letter function. 
from openpyxl.utils import get_column_letter
# Convert integer index 13 to column letter
column_letter = get_column_letter(13)
# Print or use the retrieved value
print("Column Index 13:", column_letter)
# In this example, get_column_letter(13) returns the column letter corresponding to the integer index 13.

# To get the string name for column 14 in openpyxl, you can use the openpyxl.utils.get_column_letter function.
from openpyxl.utils import get_column_letter
# Convert integer index 14 to column letter
column_letter = get_column_letter(14)
# Print or use the retrieved value
print("Column Index 14:", column_letter)
# In this example, get_column_letter(14) returns the column letter corresponding to the integer index 14. 

# To retrieve a tuple of all the Cell objects from A1 to F1 in openpyxl, you can use the following code:
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example.xlsx')
# Specify the worksheet (e.g., 'Sheet1')
worksheet = workbook['Sheet1']
# Define the range from A1 to F1
cell_range = worksheet['A1:F1']
# Convert the cell range to a tuple of Cell objects
cell_tuple = tuple(cell_range)
# Print or use the retrieved tuple of Cell objects
print("Tuple of Cell objects:", cell_tuple)
# Close the workbook when you're done
workbook.close()
# In this example, worksheet['A1:F1'] specifies the range of cells from A1 to F1. 
# The tuple(cell_range) then converts this range into a tuple of Cell objects. 
# Each element of the tuple corresponds to a Cell object in the specified range.

# To save the workbook to the filename 'example.xlsx' in openpyxl, you can use the save method of the Workbook object.
from openpyxl import Workbook
# Create a new workbook
workbook = Workbook()
# Do operations on the workbook (e.g., add data, formatting, etc.)
# Save the workbook to the filename 'example.xlsx'
workbook.save('example.xlsx')
# Close the workbook when you're done
workbook.close()
# In this example, the workbook.save('example.xlsx') line saves the workbook to the file named 'example.xlsx'.

# To set a formula in a cell using openpyxl, you can assign the formula as a string to the value attribute of the cell. 
from openpyxl import Workbook
# Create a new workbook
workbook = Workbook()
# Select the active worksheet
worksheet = workbook.active
# Set a formula in cell A1
formula = "=SUM(B1:B3)"
worksheet['A1'].value = formula
# Save the workbook
workbook.save('example_with_formula.xlsx')
# Close the workbook when you're done
workbook.close()
# In this example, the formula "=SUM(B1:B3)" is set in cell A1 by assigning it to worksheet['A1'].value. 
# When you open the resulting Excel file, you will see that cell A1 contains the result of the formula.

# If you want to retrieve the result of a cell's formula instead of the formula itself in openpyxl, you need to ensure that the workbook is in a calculated state. 
# By default, openpyxl doesn't automatically calculate the formula results when you access cell values. 
# Instead, you need to manually trigger the calculation before retrieving the result.
from openpyxl import load_workbook
# Load an existing workbook
workbook = load_workbook('example_with_formula.xlsx')
# Select the active worksheet
worksheet = workbook.active
# Manually trigger the calculation
workbook.calculation = 'automatic'  # or 'manual' if needed
workbook.save('example_with_formula.xlsx')  # Save the workbook after setting calculation mode
# Retrieve the result of the formula in cell A1
result = worksheet['A1'].value
# Print or use the retrieved result
print("Result of the formula in A1:", result)
# Close the workbook when you're done
workbook.close()
# In this example, setting workbook.calculation to 'automatic' triggers the calculation of formulas in the workbook. 
# After setting the calculation mode, you can retrieve the result of the formula using worksheet['A1'].value.

# To set the height of row 5 to 100 in openpyxl, you can use the row_dimensions attribute of the worksheet along with the height property. 
from openpyxl import Workbook
# Create a new workbook
workbook = Workbook()
# Select the active worksheet
worksheet = workbook.active
# Set the height of row 5 to 100
worksheet.row_dimensions[5].height = 100
# Save the workbook
workbook.save('example_set_row_height.xlsx')
# Close the workbook when you're done
workbook.close()
# In this example, worksheet.row_dimensions[5].height = 100 sets the height of row 5 to 100. 

# To hide column C in openpyxl, you can use the column_dimensions attribute of the worksheet along with the hidden property.
from openpyxl import Workbook
# Create a new workbook
workbook = Workbook()
# Select the active worksheet
worksheet = workbook.active
# Hide column C
worksheet.column_dimensions['C'].hidden = True
# Save the workbook
workbook.save('example_hide_column_C.xlsx')
# Close the workbook when you're done
workbook.close()
# In this example, worksheet.column_dimensions['C'].hidden = True hides column C.

# The following are a few features that openpyxl might not handle or load:
# Macros: OpenPyXL does not support the extraction or execution of macros in Excel files.
# Charts: While openpyxl allows you to manipulate and create basic charts, it may not fully support loading or rendering all types of complex charts with intricate settings.
# VBA (Visual Basic for Applications): OpenPyXL does not handle VBA code or scripts embedded in Excel files.
# Form Controls: Certain form controls, especially those associated with macros or VBA, may not be fully supported.
# Cell Comments: Although openpyxl provides support for reading and writing cell comments, there might be some advanced features or formatting that are not preserved.
# Conditional Formatting: While openpyxl supports basic conditional formatting, some advanced rules or formatting options may not be fully loaded or preserved.
# Data Validation Rules: OpenPyXL may not handle all types of data validation rules and criteria.
# Hyperlinks: There might be certain advanced hyperlink features or settings that openpyxl does not fully support.

# A freeze pane is a feature in spreadsheet applications, including Microsoft Excel, that allows you to lock or freeze a specific area of a worksheet so that it remains visible while you scroll through the rest of the sheet. 

# Creating a bar chart in openpyxl involves several steps, including creating a chart object, specifying data, and adding the chart to a worksheet. 
# Here are five functions and methods you might need to call to create a bar chart using openpyxl:
# Create a Workbook and Worksheet:
from openpyxl import Workbook
# Create a new workbook
workbook = Workbook()
# Get the active worksheet
worksheet = workbook.active
# Prepare Data:
# Example data
data = [
    ("Category", "Value"),
    ("A", 10),
    ("B", 20),
    ("C", 15),
]
# Create a BarChart object:
from openpyxl.chart import BarChart, Reference
# Create a bar chart
chart = BarChart()
# Add Data to the Chart:
# Add data to the chart
data_range = Reference(worksheet, min_col=2, min_row=1, max_col=2, max_row=len(data))
categories_range = Reference(worksheet, min_col=1, min_row=2, max_row=len(data))
chart.add_data(data_range, titles_from_data=True)
chart.set_categories(categories_range)
# Add the Chart to the Worksheet:
# Add the chart to the worksheet
worksheet.add_chart(chart, "E5")




