import openpyxl

# Open the existing workbook
workbook = openpyxl.load_workbook('output_spreadsheet.xlsx')

# Select the active sheet
sheet = workbook.active

# Specify the output text file names
output_files = ['output_column_A.txt', 'output_column_B.txt', 'output_column_C.txt']

# Loop through columns and write content to corresponding text files
for col_index, output_file in enumerate(output_files, start=1):
    with open(output_file, 'w') as file:
        # Loop through rows in the column and write to the text file
        for row_index in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_index, column=col_index).value
            if cell_value is not None:
                file.write(str(cell_value) + '\n')

# Close the workbook
workbook.close()
