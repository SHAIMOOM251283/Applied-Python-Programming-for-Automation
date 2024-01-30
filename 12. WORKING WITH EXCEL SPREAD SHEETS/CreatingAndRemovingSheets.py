import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames)
print(wb.create_sheet())
print(wb.sheetnames)
print(wb.create_sheet(index=0, title='First Sheet'))
print(wb.sheetnames)
print(wb.create_sheet(index=2, title='Middle Sheet'))
print(wb.sheetnames)

print(wb.sheetnames)
wb.remove(wb['Middle Sheet'])
wb.remove(wb['Sheet1'])
print(wb.sheetnames)
